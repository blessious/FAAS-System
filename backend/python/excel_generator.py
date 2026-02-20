import os
import sys
import mysql.connector
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
import json

# Load environment variables
load_dotenv()

class FAASExcelGenerator:
    def __init__(self):
        self.db_config = {
            'host': os.getenv('DB_HOST'),
            'user': os.getenv('DB_USER'),
            'password': os.getenv('DB_PASSWORD'),
            'database': os.getenv('DB_NAME'),
            'port': os.getenv('DB_PORT', 3306)
        }
        self.template_dir = os.path.join(os.path.dirname(__file__), 'templates')
        self.output_dir = os.path.join(os.path.dirname(__file__), 'generated')
        self.faas_dir = os.path.join(self.output_dir, 'FAAS')
        self.unirrig_dir = os.path.join(self.output_dir, 'UNIRRIG')
        os.makedirs(self.faas_dir, exist_ok=True)
        os.makedirs(self.unirrig_dir, exist_ok=True)
    
    def get_db_connection(self):
        try:
            return mysql.connector.connect(**self.db_config)
        except Exception as e:
            print(f"ERROR: Database connection failed. {e}")
            raise

    def safe_float(self, value, default=0.0):
        if value is None or value == '':
            return default
        try:
            if isinstance(value, str):
                value = value.replace(',', '')
            return float(value)
        except (ValueError, TypeError):
            return default

    def safe_int(self, value, default=0):
        if value is None or value == '':
            return default
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default

    def calculate_land_unit_value(self, classification, sub_class):
        if not classification:
            return None
        cls = str(classification).upper().strip()
        sub = str(sub_class).upper().strip() if sub_class else ""
        if cls == "COMMERCIAL":
            return {"C-1": 1400, "C-2": 1030, "C-3": 850, "C-4": 720}.get(sub)
        elif cls == "RESIDENTIAL":
            return {"R-1": 760, "R-2": 610, "R-3": 470, "R-4": 280, "R-5": 200}.get(sub)
        elif cls == "COCAL":
            return {"1": 93830, "2": 84500, "3": 76130}.get(sub)
        elif cls == "IRRIGATED":
            return {"1": 188760, "2": 174390, "3": 126950}.get(sub)
        elif cls == "UNIRRIGATED":
            return {"1": 82450, "2": 75070, "3": 52590}.get(sub)
        elif cls == "UPLAND":
            return 28540
        elif cls == "ORCHARD":
            return 44100
        elif cls == "COGON LAND":
            return 46460
        elif cls == "NIPA LAND":
            return 55070
        elif cls == "FOREST LAND":
            return 59660
        elif cls == "FISHPOND":
            return 218240
        return None

    def mround(self, number, multiple):
        if not number or not multiple:
            return 0
        return round(number / multiple) * multiple

    def calculate_improvement_unit_value(self, product_class):
        if not product_class:
            return None
        pc = str(product_class).upper().strip()
        mapping = {
            "COCO BRG.-1": 480, "COCO BRG.-2": 420, "COCO BRG.-3": 330,
            "AVOCADO": 310, "BANANA": 200, "CACAO": 300, "CALAMANSI": 360,
            "CAMANSI": 420, "CHICO": 630, "COFFEE": 250, "JACKFRUIT": 840,
            "LANZONES": 460, "MABOLO": 350, "MANGO": 1750, "ORANGE": 460,
            "RAMBUTAN": 360, "SANTOL": 620, "SINEGUELAS": 310,
            "STAR APPLE": 790, "TAMARIND": 350, "BAMBOO": 460,
            "BURI": 410, "NIPA": 240
        }
        return mapping.get(pc)

    def clean_filename(self, filename):
        if not filename:
            return "Unknown"
        filename = filename.replace(' ', '_').replace('.', '_')
        to_replace = ['\\', '/', ':', '*', '?', '"', '<', '>', '|', '(', ')', '[', ']', '{', '}', ',', ';', "'", '`', '@', '#', '$', '%', '^', '&', '+', '=', '!', '~']
        for char in to_replace:
            filename = filename.replace(char, '_')
        while '__' in filename:
            filename = filename.replace('__', '_')
        return filename.strip('_')
    
    def get_faas_record(self, record_id):
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor(dictionary=True)
            query = """
            SELECT f.*, ue.full_name as encoder_name, ua.full_name as approver_name,
                f.previous_td_no, f.previous_owner, f.effectivity_year, f.taxability,
                f.previous_av_land, f.previous_av_improvements,
                f.memoranda_code, f.memoranda_paragraph, f.owner_administrator
            FROM faas_records f
            LEFT JOIN users ue ON f.encoder_id = ue.id
            LEFT JOIN users ua ON f.approver_id = ua.id
            WHERE f.id = %s
            """
            cursor.execute(query, (record_id,))
            record = cursor.fetchone()
            cursor.close()
            conn.close()
            return record
        except Exception as e:
            print(f"Error fetching FAAS record: {e}")
            return None

    def get_unirrig_record(self, record_id):
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor(dictionary=True)
            query = "SELECT f.* FROM faas_records f WHERE f.id = %s"
            cursor.execute(query, (record_id,))
            record = cursor.fetchone()
            cursor.close()
            conn.close()
            return record
        except Exception as e:
            print(f"Error fetching UNIRRIG record: {e}")
            return None
    
    def safe_write_cell(self, sheet, cell_address, value):
        try:
            target_cell = sheet[cell_address]
            coord = target_cell.coordinate
            for merged_range in sheet.merged_cells.ranges:
                if coord in merged_range:
                    r, c = merged_range.min_row, merged_range.min_col
                    sheet.cell(row=r, column=c).value = value
                    return
            target_cell.value = value
        except Exception as e:
            print(f"Warning: Could not write to cell {cell_address}: {e}")
    
    def generate_faas_excel(self, record_id, timestamp=None, save_path=None):
        try:
            record = self.get_faas_record(record_id)
            if not record:
                return None, "FAAS record not found"

            template_path = os.path.join(self.template_dir, 'FAAS_Template.xlsx')
            if not os.path.exists(template_path):
                return None, f"Template not found at {template_path}"

            workbook = load_workbook(template_path)
            sheet = workbook.active

            print(f"Processing FAAS record for: {record.get('owner_name', 'Unknown')}")

            land_appraisals = json.loads(record.get('land_appraisals_json', '[]')) if record.get('land_appraisals_json') else []
            improvements = json.loads(record.get('improvements_json', '[]')) if record.get('improvements_json') else []
            market_values = json.loads(record.get('market_values_json', '[]')) if record.get('market_values_json') else []
            assessments = json.loads(record.get('assessments_json', '[]')) if record.get('assessments_json') else []

            cell_mapping = {
                'B6': record.get('arf_no', ''),
                'F6': record.get('pin', ''),
                'B7': record.get('oct_tct_no', ''),
                'F7': record.get('cln', ''),
                'B8': record.get('owner_name', ''),
                'F8': record.get('owner_address', ''),
                'B10': record.get('administrator_name', ''),
                'F10': record.get('administrator_address', ''),
                'B13': record.get('property_location', ''),
                'F13': record.get('property_municipality', ''),
                'B14': record.get('property_barangay', ''),
                'F14': record.get('property_province', ''),
                'B16': record.get('north_boundary', ''),
                'B17': record.get('east_boundary', ''),
                'B18': record.get('west_boundary', ''),
                'B19': record.get('south_boundary', ''),
                'B65': record.get('previous_td_no', ''),
                'B66': record.get('previous_owner', ''),
                'E65': record.get('effectivity_year', ''),
                'E66': str(record.get('taxability', '')).upper() if record.get('taxability') else '',
                'H65': self.safe_float(record.get('previous_av_land'), default=None),
                'H66': self.safe_float(record.get('previous_av_improvements'), default=None),
                'A51': record.get('owner_administrator', ''),
                'B59': str(record.get('memoranda_code', '')).upper() if record.get('memoranda_code') else '',
                'A60': record.get('memoranda_paragraph', ''),
            }

            for i in range(min(4, len(land_appraisals))):
                item = land_appraisals[i]
                row = 22 + i
                classification = item.get('classification')
                sub_class = item.get('sub_class')
                if classification:
                    cell_mapping[f'A{row}'] = classification
                if sub_class:
                    cell_mapping[f'C{row}'] = sub_class
                if item.get('area'):
                    cell_mapping[f'D{row}'] = self.safe_float(item.get('area'))
                unit_value = self.calculate_land_unit_value(classification, sub_class)
                area = self.safe_float(item.get('area'))
                if unit_value:
                    cell_mapping[f'E{row}'] = unit_value
                    if area > 0:
                        cell_mapping[f'G{row}'] = self.mround(area * unit_value, 10)

            for i in range(min(4, len(improvements))):
                item = improvements[i]
                row = 29 + i
                product_class = item.get('product_class')
                qty = self.safe_int(item.get('improvement_qty'))
                if product_class:
                    cell_mapping[f'A{row}'] = product_class
                if qty:
                    cell_mapping[f'C{row}'] = qty
                unit_value = self.calculate_improvement_unit_value(product_class)
                if unit_value is None and item.get('unit_value_improvement'):
                    unit_value = self.safe_float(item.get('unit_value_improvement'))
                if unit_value:
                    cell_mapping[f'E{row}'] = unit_value
                    if qty > 0:
                        cell_mapping[f'G{row}'] = qty * unit_value

            # Market Value rows 36-39: land first, then improvements, no gaps
            mv_sources: list[float] = []
            for i in range(min(4, len(land_appraisals))):
                mv = cell_mapping.get(f'G{22 + i}', 0)
                if mv:
                    mv_sources.append(mv)
            for i in range(min(4, len(improvements))):
                mv = cell_mapping.get(f'G{29 + i}', 0)
                if mv:
                    mv_sources.append(mv)

            for i, mv in enumerate(mv_sources[:4]):
                row = 36 + i
                mv_item = market_values[i] if i < len(market_values) else {}
                adj_factor = mv_item.get('adj_factor', '')
                percent_adj = self.safe_float(mv_item.get('percent_adjustment'))
                cell_mapping[f'A{row}'] = mv
                if adj_factor:
                    cell_mapping[f'C{row}'] = adj_factor
                if percent_adj:
                    cell_mapping[f'D{row}'] = percent_adj / 100.0
                    calc_g = self.mround(mv * (percent_adj / 100.0), 10)
                    cell_mapping[f'G{row}'] = calc_g if calc_g != 0 else ''
                else:
                    cell_mapping[f'G{row}'] = mv

            # Assessment rows 43-46
            g36_values = [cell_mapping.get(f'G{36 + i}', 0) for i in range(4)]
            for i in range(min(4, len(assessments))):
                item = assessments[i]
                row = 43 + i
                assessment_level = self.safe_float(item.get('assessment_level'))
                if item.get('kind'):
                    cell_mapping[f'A{row}'] = item.get('kind')
                if item.get('actual_use'):
                    cell_mapping[f'B{row}'] = item.get('actual_use')
                mv_from_g = g36_values[i]
                if mv_from_g:
                    cell_mapping[f'C{row}'] = mv_from_g
                if assessment_level:
                    cell_mapping[f'E{row}'] = assessment_level / 100.0
                if mv_from_g and assessment_level:
                    calc_av = self.mround(mv_from_g * (assessment_level / 100.0), 10)
                    cell_mapping[f'G{row}'] = calc_av if calc_av != 0 else ""

            for cell, value in cell_mapping.items():
                if value != '':
                    if isinstance(value, str) and value.lower() == 'none':
                        continue
                    self.safe_write_cell(sheet, cell, value)

            owner_raw = record.get('owner_name', 'Unknown')
            owner_name = self.clean_filename(owner_raw)
            arf_raw = record.get('arf_no', 'Unknown')
            arf_no_safe = self.clean_filename(arf_raw)
            if not timestamp:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"FAAS_{arf_no_safe}_{owner_name}_{timestamp}.xlsx"
            output_path = save_path if save_path else os.path.join(self.faas_dir, filename)
            workbook.save(output_path)
            print(f"[OK] Excel file generated: {output_path}")
            print(f"[INFO] Record details: ARF={record.get('arf_no', 'N/A')}, Owner={record.get('owner_name', 'N/A')}")
            return {'success': True, 'file_path': output_path, 'file_name': filename, 'owner_name': owner_name, 'arf_no': record.get('arf_no', '')}, None

        except Exception as e:
            import traceback
            error_msg = f"Error generating Excel: {str(e)}\n{traceback.format_exc()}"
            print(f" {error_msg}")
            return None, error_msg

    def generate_unirrig_excel(self, record_id, timestamp=None, save_path=None):
        try:
            record = self.get_unirrig_record(record_id)
            if not record:
                return None, "UNIRRIG record not found"

            template_path = os.path.join(self.template_dir, 'UNIRRIG_Template.xlsx')
            if not os.path.exists(template_path):
                return None, f"Template not found at {template_path}"

            workbook = load_workbook(template_path)
            sheet = workbook['Sheet1']  # always use Sheet1 by name

            # Set page margins for Sheet1
            sheet.page_margins.top = 0.25
            sheet.page_margins.bottom = 0.25

            print(f"Processing UNIRRIG record for: {record.get('owner_name', 'Unknown')}")

            # Debug: print sample field values

            land_appraisals = json.loads(record.get('land_appraisals_json', '[]')) if record.get('land_appraisals_json') else []
            improvements = json.loads(record.get('improvements_json', '[]')) if record.get('improvements_json') else []
            market_values = json.loads(record.get('market_values_json', '[]')) if record.get('market_values_json') else []
            assessments = json.loads(record.get('assessments_json', '[]')) if record.get('assessments_json') else []


            # ── SHEET 1 MAPPING ──
            unirrig_mapping = {
                'J3':  record.get('pin', ''),
                'A7':  record.get('memoranda_code', ''),
                'B11': record.get('owner_name', ''),
                'H11': record.get('owner_address', ''),
                'B13': record.get('administrator_name', ''),
                'H13': record.get('administrator_address', ''),
                'C17': record.get('property_location', ''),
                'I17': record.get('property_municipality', ''),
                'G17': record.get('property_province', ''),
                'C19': record.get('oct_tct_no', ''),
                'H19': record.get('cln', ''),
                'B21': record.get('north_boundary', ''),
                'I21': record.get('south_boundary', ''),
                'B22': record.get('east_boundary', ''),
                'I22': record.get('west_boundary', ''),
                'H62': record.get('memoranda_paragraph', ''),
            }

            # % Adjustment → G50
            if market_values:
                percent_adj = market_values[0].get('percent_adjustment')
                if percent_adj:
                    unirrig_mapping['G50'] = self.safe_float(percent_adj) / 100.0

            # Adj. Factor parts → G42, G45, G48
            if market_values:
                adj_factor_raw = market_values[0].get('adj_factor', '')
                if adj_factor_raw:
                    adj_parts = [p.strip() for p in str(adj_factor_raw).split(',')]
                    if len(adj_parts) > 0:
                        unirrig_mapping['G42'] = self.safe_float(adj_parts[0]) / 100.0
                    if len(adj_parts) > 1:
                        unirrig_mapping['G45'] = self.safe_float(adj_parts[1]) / 100.0
                    if len(adj_parts) > 2:
                        unirrig_mapping['G48'] = self.safe_float(adj_parts[2]) / 100.0

            # Write Sheet1 mapping
            for cell, value in unirrig_mapping.items():
                if value != '' and value is not None:
                    self.safe_write_cell(sheet, cell, value)

            # PIN → K19, K20
            pin = record.get('pin', '')
            pin_parts = pin.split('-')
            if len(pin_parts) >= 5:
                self.safe_write_cell(sheet, 'K20', pin_parts[3])
                self.safe_write_cell(sheet, 'K19', pin_parts[4])

            # Land Appraisals → rows 28-31
            for i in range(min(4, len(land_appraisals))):
                item = land_appraisals[i]
                unirrig_row = 28 + i
                classification = item.get('classification')
                sub_class = item.get('sub_class')
                if classification:
                    self.safe_write_cell(sheet, f'E{unirrig_row}', classification)
                if sub_class:
                    self.safe_write_cell(sheet, f'H{unirrig_row}', sub_class)
                if item.get('area'):
                    self.safe_write_cell(sheet, f'G{unirrig_row}', f"{self.safe_float(item.get('area')):,.4f}")
                unit_value = self.calculate_land_unit_value(classification, sub_class)
                if unit_value:
                    self.safe_write_cell(sheet, f'I{unirrig_row}', unit_value)

            # ── COMPUTE ADJUSTED MARKET VALUES (MATCHES FAAS G36-G39) ──
            base_mvs: list[float] = []
            for item in land_appraisals:
                cls = item.get('classification', '')
                sub = item.get('sub_class', '')
                area = self.safe_float(item.get('area'))
                unit_val = self.calculate_land_unit_value(cls, sub) or 0
                mv = self.mround(area * unit_val, 10) if area and unit_val else 0
                if mv: base_mvs.append(float(mv))
            
            for i, item in enumerate(improvements):
                pc = item.get('product_class', '')
                qty = self.safe_int(item.get('improvement_qty'))
                u_val = self.calculate_improvement_unit_value(pc)
                if u_val is None and item.get('unit_value_improvement'):
                    u_val = self.safe_float(item.get('unit_value_improvement'))
                u_val_safe = u_val or 0
                mv = self.mround(qty * u_val_safe, 10) if qty and u_val_safe else 0
                if mv: base_mvs.append(float(mv))

                # Write Improvements to Sheet 1 → rows 42-45
                if i < 4:
                    unirrig_row = 42 + i
                    if pc:
                        self.safe_write_cell(sheet, f'H{unirrig_row}', pc)
                    if qty:
                        self.safe_write_cell(sheet, f'I{unirrig_row}', qty)
                    if u_val_safe:
                        self.safe_write_cell(sheet, f'J{unirrig_row}', u_val_safe)
                    if mv:
                        self.safe_write_cell(sheet, f'K{unirrig_row}', mv)

            adjusted_mvs: list[float] = []
            for i in range(min(4, len(base_mvs))):
                mv = base_mvs[i]
                mv_item = market_values[i] if i < len(market_values) else {}
                percent_adj = self.safe_float(mv_item.get('percent_adjustment'))
                if percent_adj:
                    calc_g = self.mround(mv * (percent_adj / 100.0), 10)
                    adjusted_mvs.append(float(calc_g) if calc_g != 0 else float(mv))
                else:
                    adjusted_mvs.append(float(mv))

            if 'Sheet2' in workbook.sheetnames:
                sheet2 = workbook['Sheet2']
                sheet2.page_margins.top = 0.25
                sheet2.page_margins.bottom = 0.25

                # J36 result: Sum of Land MVs * % Adjustment
                total_land_mv = 0.0
                for item in land_appraisals:
                    cls = item.get('classification', '')
                    sub = item.get('sub_class', '')
                    area = self.safe_float(item.get('area'))
                    u_val = self.calculate_land_unit_value(cls, sub) or 0
                    mv = self.mround(area * u_val, 10) if area and u_val else 0
                    total_land_mv += float(mv)
                
                pct_adj = 0.0
                if market_values:
                    pct_adj = self.safe_float(market_values[0].get('percent_adjustment')) / 100.0
                
                if pct_adj != 0:
                    j36_final = self.mround(total_land_mv * pct_adj, 10)
                    self.safe_write_cell(sheet2, 'J36', j36_final)
                else:
                    self.safe_write_cell(sheet2, 'J36', total_land_mv)

                # J37 result: Sum of Improvements MVs * % Adjustment
                total_impr_mv = 0.0
                for item in improvements:
                    pc = item.get('product_class', '')
                    qty = self.safe_int(item.get('improvement_qty'))
                    u_val = self.calculate_improvement_unit_value(pc)
                    if u_val is None and item.get('unit_value_improvement'):
                        u_val = self.safe_float(item.get('unit_value_improvement'))
                    u_val_safe = u_val or 0
                    mv = self.mround(qty * u_val_safe, 10) if qty and u_val_safe else 0
                    total_impr_mv += float(mv)

                if pct_adj != 0:
                    j37_final = self.mround(total_impr_mv * pct_adj, 10)
                    self.safe_write_cell(sheet2, 'J37', j37_final)
                else:
                    self.safe_write_cell(sheet2, 'J37', total_impr_mv)

                for i in range(min(4, len(assessments))):
                    item = assessments[i]
                    row = 54 + i
                    kind = item.get('kind', '')
                    actual_use = item.get('actual_use', '')
                    assessment_level = self.safe_float(item.get('assessment_level'))
                    market_val_detail = adjusted_mvs[i] if i < len(adjusted_mvs) else 0

                    if kind:
                        self.safe_write_cell(sheet2, f'A{row}', kind)
                    if actual_use:
                        self.safe_write_cell(sheet2, f'D{row}', actual_use)
                    if market_val_detail:
                        self.safe_write_cell(sheet2, f'G{row}', market_val_detail)
                    if assessment_level:
                        self.safe_write_cell(sheet2, f'I{row}', assessment_level / 100.0)
                    if market_val_detail and assessment_level:
                        calc_av = self.mround(market_val_detail * (assessment_level / 100.0), 10)
                        self.safe_write_cell(sheet2, f'K{row}', calc_av if calc_av != 0 else '')

                print(f"[OK] UNIRRIG Page 2 assessments written: {len(assessments)} row(s)")
            else:
                print("[WARN] UNIRRIG template has only 1 sheet — add Sheet2 to UNIRRIG_Template.xlsx first.")

            owner_raw = record.get('owner_name', 'Unknown')
            owner_name_safe = self.clean_filename(owner_raw)
            arf_raw = record.get('arf_no', 'Unknown')
            arf_no_safe = self.clean_filename(arf_raw)
            if not timestamp:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"UNIRRIG_{arf_no_safe}_{owner_name_safe}_{timestamp}.xlsx"
            output_path = save_path if save_path else os.path.join(self.unirrig_dir, filename)
            workbook.save(output_path)
            print(f"[OK] UNIRRIG Excel file generated: {output_path}")
            return {'success': True, 'file_path': output_path, 'file_name': filename, 'owner_name': owner_name_safe}, None

        except Exception as e:
            import traceback
            error_msg = f"Error generating UNIRRIG Excel: {str(e)}\n{traceback.format_exc()}"
            print(f" {error_msg}")
            return None, error_msg

    def inspect_template(self):
        try:
            template_path = os.path.join(self.template_dir, 'FAAS_Template.xlsx')
            workbook = load_workbook(template_path)
            sheet = workbook.active
            print(f"Sheet: {sheet.title}, Max row: {sheet.max_row}, Max col: {sheet.max_column}")
            workbook.close()
        except Exception as e:
            print(f"Error inspecting template: {e}")

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--record-id', type=int, required=True)
    parser.add_argument('--type', type=str, choices=['faas', 'unirrig', 'both'], default='faas')
    args = parser.parse_args()
    generator = FAASExcelGenerator()

    if args.type == 'faas':
        result, error = generator.generate_faas_excel(args.record_id)
        print(json.dumps(result if result else {"success": False, "error": error}))
        sys.exit(0 if result else 1)
    elif args.type == 'unirrig':
        result, error = generator.generate_unirrig_excel(args.record_id)
        print(json.dumps(result if result else {"success": False, "error": error}))
        sys.exit(0 if result else 1)
    elif args.type == 'both':
        shared_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        faas_result, faas_error = generator.generate_faas_excel(args.record_id, timestamp=shared_timestamp)
        unirrig_result, unirrig_error = generator.generate_unirrig_excel(args.record_id, timestamp=shared_timestamp)
        is_faas_success = bool(faas_result)
        output = {
            "success": is_faas_success,
            "both_success": is_faas_success and bool(unirrig_result),
            "faas": faas_result,
            "unirrig": unirrig_result,
            "faas_error": faas_error if not faas_result else None,
            "unirrig_error": unirrig_error if not unirrig_result else None
        }
        print(json.dumps(output))
        sys.exit(0 if is_faas_success else 1)
    else:
        print(json.dumps({"success": False, "error": "Invalid type"}))
        sys.exit(1)

if __name__ == "__main__":
    main()