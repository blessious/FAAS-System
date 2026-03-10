import os
import sys
import json
import argparse
from reportlab.lib.units import cm, inch
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook

# 🎯 MASTER TEMPLATE - Default coordinates (CM from Bottom-Left)
TEMPLATE_MAPPING = {
    'Sheet1!J3':  {'x': 16.0, 'y': 26.5, 'label': 'PIN'},
    'Sheet1!A7':  {'x': 2.5,  'y': 24.5, 'label': 'Memoranda'},
    'Sheet1!B11': {'x': 4.0,  'y': 23.3, 'label': 'Owner'},
    'Sheet1!B13': {'x': 4.3,  'y': 22.75, 'label': 'Admin'},
    'Sheet1!H11': {'x': 12.0, 'y': 23.3, 'label': 'Owner Addr'},
    'Sheet1!H13': {'x': 12.0, 'y': 22.75, 'label': 'Admin Addr'},
    'Sheet1!C17': {'x': 7.0,  'y': 21.5, 'label': 'Street'},
    'Sheet1!G17': {'x': 10.5, 'y': 21.5, 'label': 'Barangay'},
    'Sheet1!I17': {'x': 16.0, 'y': 21.5, 'label': 'Municipality'},
    'Sheet1!C19': {'x': 5.5,  'y': 20.7, 'label': 'Title'},
    'Sheet1!H19': {'x': 11.8, 'y': 20.6, 'label': 'Lot No'},
    'Sheet1!K19': {'x': 18.0, 'y': 20.65, 'label': 'K19'},
    'Sheet1!K20': {'x': 18.0, 'y': 20.2, 'label': 'K20'},
    'Sheet1!B21': {'x': 4.0,  'y': 20.5, 'label': 'North'},
    'Sheet1!B22': {'x': 4.0,  'y': 19.4, 'label': 'East'},
    'Sheet1!I21': {'x': 12.3, 'y': 19.75, 'label': 'South'},
    'Sheet1!I22': {'x': 12.3, 'y': 19.4, 'label': 'West'},
    'Sheet1!E28': {'x': 9.6,  'y': 16.8,  'label': 'Table R1 Class'},
    'Sheet1!G28': {'x': 11.6, 'y': 16.8,  'label': 'Table R1 Area'},
    'Sheet1!H28': {'x': 14.0, 'y': 16.8,  'label': 'Table R1 Sub'},
    'Sheet1!I28': {'x': 15.5, 'y': 16.8,  'label': 'Table R1 Unit'},
    'Sheet1!J28': {'x': 17.6, 'y': 16.8,  'label': 'Table R1 Market'},
    'Sheet1!E33': {'x': 8.2,  'y': 15.95, 'label': 'Table R6 Class'},
    'Sheet1!G33': {'x': 11.0, 'y': 15.45, 'label': 'Table R6 Area'},
    'Sheet1!H33': {'x': 13.0, 'y': 14.95, 'label': 'Table R6 Sub'},
    'Sheet1!I33': {'x': 15.1, 'y': 14.35, 'label': 'Table R6 Unit'},
    'Sheet1!J33': {'x': 17.5, 'y': 13.85, 'label': 'Table R6 Market'},
    'Sheet1!E35': {'x': 8.5,  'y': 13.35, 'label': 'Table R8 Class'},
    'Sheet1!G35': {'x': 11.0, 'y': 12.85, 'label': 'Table R8 Area'},
    'Sheet1!H35': {'x': 13.0, 'y': 12.35, 'label': 'Table R8 Sub'},
    'Sheet1!I35': {'x': 15.0, 'y': 11.85, 'label': 'Table R8 Unit'},
    'Sheet1!J35': {'x': 17.5, 'y': 11.35, 'label': 'Table R8 Market'},
    'Sheet1!J36': {'x': 17.8, 'y': 13.75, 'label': 'Total MV'},
    'Sheet1!H42': {'x': 12.7, 'y': 11.8,  'label': 'Plant: Kind'},
    'Sheet1!I42': {'x': 15.0, 'y': 11.8,  'label': 'Plant: Area'},
    'Sheet1!J42': {'x': 17.0, 'y': 11.8,  'label': 'Plant: UV'},
    'Sheet1!K42': {'x': 18.2, 'y': 11.8,  'label': 'Plant: MV'},
    'Sheet1!K53': {'x': 18.4, 'y': 7.8,   'label': 'Total MV Plants'},
    'Sheet1!G44': {'x': 11.0, 'y': 11.4,  'label': 'Plant Adj 1 %'},
    'Sheet1!G47': {'x': 11.0, 'y': 10.6,  'label': 'Plant Adj 2 %'},
    'Sheet1!G49': {'x': 11.0, 'y': 9.7,   'label': 'Plant Adj 3 %'},
    'Sheet1!G52': {'x': 11.0, 'y': 9.3,   'label': 'Plant Adj 4 %'},
    'Sheet1!E58': {'x': 8.7,  'y': 5.2,   'label': 'Land II: Kind'},
    'Sheet1!G58': {'x': 11.3, 'y': 5.2,   'label': 'Land II: Area'},
    'Sheet1!H58': {'x': 13.7, 'y': 5.2,   'label': 'Land II: Class'},
    'Sheet1!H59': {'x': 13.7, 'y': 4.5,   'label': 'Land II: UV'},
    'Sheet1!J58': {'x': 17.6, 'y': 5.1,   'label': 'Land II: MV'},
    # Additional table rows (R2-R4)
    'Sheet1!E29': {'x': 8.8, 'y': 16.35, 'label': 'Table R2 Class'},
    'Sheet1!G29': {'x': 11.5,'y': 16.35, 'label': 'Table R2 Area'},
    'Sheet1!H29': {'x': 13.5,'y': 16.3,  'label': 'Table R2 Sub'},
    'Sheet1!I29': {'x': 15.5,'y': 16.3,  'label': 'Table R2 Unit'},
    'Sheet1!J29': {'x': 17.6,'y': 16.3,  'label': 'Table R2 Market'},
    'Sheet1!E30': {'x': 8.8, 'y': 15.95, 'label': 'Table R3 Class'},
    'Sheet1!G30': {'x': 11.5,'y': 15.95, 'label': 'Table R3 Area'},
    'Sheet1!H30': {'x': 13.5,'y': 15.9,  'label': 'Table R3 Sub'},
    'Sheet1!I30': {'x': 15.5,'y': 15.9,  'label': 'Table R3 Unit'},
    'Sheet1!J30': {'x': 17.6,'y': 15.9,  'label': 'Table R3 Market'},
    'Sheet1!E31': {'x': 8.8, 'y': 15.5,  'label': 'Table R4 Class'},
    'Sheet1!G31': {'x': 11.5,'y': 15.5,  'label': 'Table R4 Area'},
    'Sheet1!H31': {'x': 13.5,'y': 15.5,  'label': 'Table R4 Sub'},
    'Sheet1!I31': {'x': 15.5,'y': 15.5,  'label': 'Table R4 Unit'},
    'Sheet1!J31': {'x': 17.6,'y': 15.5,  'label': 'Table R4 Market'},
    'Sheet2!L36': {'x': 14.1, 'y': 14.1, 'label': 'Total Land MV'},
    'Sheet2!L37': {'x': 14.1, 'y': 14.3, 'label': 'Total Impr MV'},
    'Sheet2!L38': {'x': 14.1, 'y': 14.7, 'label': 'L38'},
    'Sheet2!E42': {'x': 9.2,  'y': 11.4, 'label': 'Sworn Day'},
    'Sheet2!G42': {'x': 11.9, 'y': 11.4, 'label': 'Sworn Month'},
    'Sheet2!I42': {'x': 16.2, 'y': 11.4, 'label': 'Sworn Year'},
    'Sheet2!G43': {'x': 11.3, 'y': 11.0, 'label': 'CTC No'},
    'Sheet2!J43': {'x': 15.9, 'y': 11.0, 'label': 'CTC Month'},
    'Sheet2!B44': {'x': 2.4,  'y': 10.7, 'label': 'CTC Year'},
    'Sheet2!D44': {'x': 3.6,  'y': 10.7, 'label': 'CTC Issued At'},
    'Sheet2!A54': {'x': 2.1,  'y': 6.9,  'label': 'Assm 1 Kind'},
    'Sheet2!A55': {'x': 2.1,  'y': 6.5,  'label': 'Assm 2 Kind'},
    'Sheet2!A56': {'x': 2.1,  'y': 6.1,  'label': 'Assm 3 Kind'},
    'Sheet2!A57': {'x': 2.1,  'y': 5.7,  'label': 'Assm 4 Kind'},
    'Sheet2!D54': {'x': 5.4,  'y': 6.9,  'label': 'Assm 1 Use'},
    'Sheet2!D55': {'x': 5.4,  'y': 6.5,  'label': 'Assm 2 Use'},
    'Sheet2!D56': {'x': 5.4,  'y': 6.1,  'label': 'Assm 3 Use'},
    'Sheet2!D57': {'x': 5.4,  'y': 5.7,  'label': 'Assm 4 Use'},
    'Sheet2!G54': {'x': 9.0,  'y': 6.9,  'label': 'Assm 1 MV'},
    'Sheet2!G55': {'x': 9.0,  'y': 6.5,  'label': 'Assm 2 MV'},
    'Sheet2!G56': {'x': 9.0,  'y': 6.1,  'label': 'Assm 3 MV'},
    'Sheet2!G57': {'x': 9.0,  'y': 5.7,  'label': 'Assm 4 MV'},
    'Sheet2!G58': {'x': 9.0,  'y': 5.3,  'label': 'Total Assm MV'},
    'Sheet2!K54': {'x': 12.9, 'y': 6.9,  'label': 'Assm 1 Lvl'},
    'Sheet2!K55': {'x': 12.9, 'y': 6.5,  'label': 'Assm 2 Lvl'},
    'Sheet2!K56': {'x': 12.9, 'y': 6.1,  'label': 'Assm 3 Lvl'},
    'Sheet2!K57': {'x': 12.9, 'y': 5.7,  'label': 'Assm 4 Lvl'},
    'Sheet2!M54': {'x': 16.2, 'y': 6.9,  'label': 'Assm 1 AV'},
    'Sheet2!M55': {'x': 16.2, 'y': 6.5,  'label': 'Assm 2 AV'},
    'Sheet2!M56': {'x': 16.2, 'y': 6.1,  'label': 'Assm 3 AV'},
    'Sheet2!M57': {'x': 16.2, 'y': 5.7,  'label': 'Assm 4 AV'},
    'Sheet2!M58': {'x': 16.2, 'y': 5.3,  'label': 'Total Assm AV'},
    'Sheet2!D59': {'x': 9.2,  'y': 5.3,  'label': 'Words'},
    'Sheet2!E67': {'x': 10.8, 'y': 2.3,  'label': 'Prev TD'},
    'Sheet2!B69': {'x': 2.8,  'y': 1.6,  'label': 'Eff Year'},
    'Sheet2!E69': {'x': 7.7,  'y': 1.6,  'label': 'Prev Owner'},
    'Sheet2!E70': {'x': 8.0,  'y': 0.8,  'label': 'Prev Land AV'},
    'Sheet2!H70': {'x': 13.3, 'y': 0.5,  'label': 'Prev Impr AV'},
    'Sheet2!L71': {'x': 16.0, 'y': 0.5,  'label': 'Total Prev AV'},
    'Sheet2!L39': {'x': 14.1, 'y': 15.0, 'label': 'Owner/Administrator'},
    'Sheet2!K39': {'x': 13.0, 'y': 12.6, 'label': 'K39'},
    'Sheet2!J45': {'x': 12.7, 'y': 10.3, 'label': 'J45'},
    'Sheet2!J47': {'x': 13.6, 'y': 9.6,  'label': 'J47'},
    'Sheet2!A61': {'x': 2.8,  'y': 3.7,  'label': 'A61'},
}

class PrecisionPDFGenerator:
    def __init__(self, mapping_file=None):
        self.font_name = "Helvetica-Bold"
        self._register_font()
        self.mapping = TEMPLATE_MAPPING.copy()
        
        if mapping_file:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            fpath = mapping_file if os.path.isabs(mapping_file) else os.path.join(script_dir, mapping_file)
            if os.path.exists(fpath):
                try:
                    with open(fpath, 'r') as f:
                        self.mapping.update(json.load(f))
                except: pass

        self.paper_size = (8 * inch, 11 * inch)

    def _register_font(self):
        font_path = r"C:\Windows\Fonts\bahnschrift.ttf"
        if os.path.exists(font_path):
            try:
                name = "Bahnschrift-Bold"
                pdfmetrics.registerFont(TTFont(name, font_path))
                self.font_name = name
            except: pass

    def generate(self, excel_path, output_path=None):
        if not os.path.exists(excel_path):
            return {"success": False, "error": "Excel missing"}
        
        if not output_path:
            out_dir = os.path.join(os.path.dirname(__file__), "generated", "PRECISION")
            os.makedirs(out_dir, exist_ok=True)
            output_path = os.path.join(out_dir, f"Precision_{os.path.basename(excel_path).replace('.xlsx', '.pdf')}")

        try:
            wb = load_workbook(excel_path, data_only=True)
            c = canvas.Canvas(output_path, pagesize=self.paper_size)
            
            fields_processed = 0
            current_sheet = None
            
            # Sort mapping by sheet name to group fields together (Sheet1 then Sheet2)
            sorted_mapping = dict(sorted(self.mapping.items(), key=lambda item: item[0]))
            
            for key, coord in sorted_mapping.items():
                try:
                    text = coord.get('text')
                    sn, addr = None, None
                    if '!' in key:
                        sn, addr = key.split('!')
                        
                        # Trigger Page 2 (Sheet 2) only when we first encounter a Sheet 2 field that HAS data
                        if current_sheet == 'Sheet1' and sn == 'Sheet2':
                            # Check if Sheet 2 actually exists and has at least one valid value to print
                            # to avoid creating a blank second page
                            has_sheet2_data = False
                            for k2, v2 in sorted_mapping.items():
                                if k2.startswith('Sheet2!'):
                                    addr2 = k2.split('!')[1]
                                    if '_line' in addr2: continue
                                    if 'Sheet2' in wb.sheetnames and wb['Sheet2'][addr2].value:
                                        has_sheet2_data = True
                                        break
                            
                            if has_sheet2_data:
                                c.showPage()
                                # Reset transformation for new page
                                c.setStrokeColorRGB(0, 0, 0)
                                c.setFillColorRGB(0, 0, 0)
                                c.setLineWidth(0.5)

                        current_sheet = sn

                        # Strip _line suffix for Excel reading
                        if '_line' in addr:
                            addr = addr.split('_line')[0]

                    if text is None:
                        if not sn: continue
                        
                        if sn not in wb.sheetnames:
                            sn_lower = sn.lower()
                            real_sn = next((s for s in wb.sheetnames if s.lower() == sn_lower), None)
                            if real_sn: sn = real_sn
                            else: continue
                        
                        # Normal Cell Reading
                        val = wb[sn][addr].value
                        if val is None or str(val).strip() == "": continue
                        
                        text = str(val)
                        # Percentage conversion for specific G cells
                        if addr in ['G44', 'G47', 'G49', 'G52', 'K54', 'K55', 'K56', 'K57']:
                            try:
                                num_val = float(val)
                                text = f"{num_val * 100:.0f}%" if num_val <= 1.0 else f"{num_val:.0f}%"
                            except: pass
                        # 🚀 No decimal places for plant Area (I42) and K19/K20
                        elif addr in ['I42', 'K19', 'K20']:
                            try: text = f"{int(float(val))}"
                            except: pass
                        # 🚀 2 decimal places for Currency/Calculated MV (I, J, K, L, M columns and G54-G58, H58)
                        elif addr and (any(col in addr for col in ['I', 'J', 'K', 'L', 'M']) or addr in ['G54', 'G55', 'G56', 'G57', 'G58', 'H58']):
                            try: text = f"{float(val):,.2f}"
                            except: pass
                    
                    if not text or str(text).strip() == "": continue
                    
                    # 🚀 Fix for J36: Calculate Total MV manually if Excel version is blank
                    if addr == 'J28': 
                        running_total = 0.0
                        for row_idx in range(28, 36):
                            val_to_add = wb[sn][f'J{row_idx}'].value
                            if val_to_add:
                                try:
                                    running_total = float(running_total) + float(val_to_add)
                                except: pass
                        
                        target_j36 = self.mapping.get('Sheet1!J36')
                        if target_j36:
                            fs_j36 = target_j36.get('fontSize', 10.5)
                            x_j36, y_j36 = target_j36['x'] * cm, target_j36['y'] * cm
                            c.saveState()
                            c.translate(x_j36, y_j36)
                            t_j = c.beginText(0, 0)
                            t_j.setFont(self.font_name, fs_j36)
                            t_j.setTextRenderMode(2)
                            c.setLineWidth(0.5)
                            t_j.setHorizScale(75)
                            t_j.textOut(f"{running_total:,.2f}")
                            c.drawText(t_j)
                            c.restoreState()
                    
                    # 🚀 G58: Sum of G54-G57
                    if addr == 'G58':
                        g58_total = 0.0
                        for row_idx in range(54, 58):
                            val_to_add = wb[sn][f'G{row_idx}'].value
                            if val_to_add:
                                try:
                                    g58_total = float(g58_total) + float(val_to_add)
                                except: pass
                        
                        if g58_total > 0:
                            target_g58 = self.mapping.get('Sheet2!G58')
                            if target_g58:
                                fs_g58 = target_g58.get('fontSize', 10.5)
                                x_g58, y_g58 = target_g58['x'] * cm, target_g58['y'] * cm
                                c.saveState()
                                c.translate(x_g58, y_g58)
                                c.scale(0.75, 1.0)
                                t_g = c.beginText(0, 0)
                                t_g.setFont(self.font_name, fs_g58)
                                t_g.setFillColorRGB(0, 0, 0)
                                t_g.setTextRenderMode(2)
                                c.setLineWidth(0.5)
                                c.setStrokeColorRGB(0, 0, 0)
                                t_g.textOut(f"{g58_total:,.2f}")
                                c.drawText(t_g)
                                c.restoreState()
                        continue  # Skip normal rendering for G58
                    
                    # 🚀 M58: Sum of M54-M57
                    if addr == 'M58':
                        m58_total = 0.0
                        for row_idx in range(54, 58):
                            val_to_add = wb[sn][f'M{row_idx}'].value
                            if val_to_add:
                                try:
                                    m58_total = float(m58_total) + float(val_to_add)
                                except: pass
                        
                        if m58_total > 0:
                            target_m58 = self.mapping.get('Sheet2!M58')
                            if target_m58:
                                fs_m58 = target_m58.get('fontSize', 10.5)
                                x_m58, y_m58 = target_m58['x'] * cm, target_m58['y'] * cm
                                c.saveState()
                                c.translate(x_m58, y_m58)
                                c.scale(0.75, 1.0)
                                t_m = c.beginText(0, 0)
                                t_m.setFont(self.font_name, fs_m58)
                                t_m.setFillColorRGB(0, 0, 0)
                                t_m.setTextRenderMode(2)
                                c.setLineWidth(0.5)
                                c.setStrokeColorRGB(0, 0, 0)
                                t_m.textOut(f"{m58_total:,.2f}")
                                c.drawText(t_m)
                                c.restoreState()
                        continue  # Skip normal rendering for M58
                    
                    # 🚀 K53: Sum of K42-K45
                    if addr == 'K53':
                        k53_total = 0.0
                        for row_idx in range(42, 46):
                            val_to_add = wb[sn][f'K{row_idx}'].value
                            if val_to_add:
                                try:
                                    k53_total = float(k53_total) + float(val_to_add)
                                except: pass
                        
                        if k53_total > 0:
                            target_k53 = self.mapping.get('Sheet1!K53')
                            if target_k53:
                                fs_k53 = target_k53.get('fontSize', 10.5)
                                x_k53, y_k53 = target_k53['x'] * cm, target_k53['y'] * cm
                                c.saveState()
                                c.translate(x_k53, y_k53)
                                c.scale(0.75, 1.0)
                                t_k = c.beginText(0, 0)
                                t_k.setFont(self.font_name, fs_k53)
                                t_k.setFillColorRGB(0, 0, 0)
                                t_k.setTextRenderMode(2)
                                c.setLineWidth(0.5)
                                c.setStrokeColorRGB(0, 0, 0)
                                t_k.textOut(f"{k53_total:,.2f}")
                                c.drawText(t_k)
                                c.restoreState()
                        continue  # Skip normal rendering for K53
                    
                    x, y = coord['x'] * cm, coord['y'] * cm

                    c.saveState()
                    # Apply transformation to the canvas
                    c.translate(x, y)
                    c.scale(0.75, 1.0)
                    
                    # Use a text object for bold rendering
                    fs = coord.get('fontSize', 10.5)
                    t = c.beginText(0, 0) # Positioned at 0,0 because of translate
                    t.setFont(self.font_name, fs)
                    t.setFillColorRGB(0, 0, 0)
                    t.setTextRenderMode(2) # Fill + Stroke
                    c.setLineWidth(0.5)
                    c.setStrokeColorRGB(0, 0, 0)
                    
                    if "Total" in str(coord.get('label','')):
                        # For Total fields, use textOut like regular fields
                        t.textOut(text)
                        c.drawText(t)
                    else:
                        t.textOut(text)
                        c.drawText(t)
                        
                    c.restoreState()
                    fields_processed += 1
                except Exception as e:
                    sys.stderr.write(f"DEBUG: Error drawing {key}: {e}\n")
                    continue

            c.save()
            return {
                "success": True, 
                "fields_count": fields_processed,
                "file_path": output_path, 
                "file_name": os.path.basename(output_path)
            }
        except Exception as ex:
            return {"success": False, "error": str(ex)}

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument('--excel-path', required=True)
    p.add_argument('--mapping-file')
    a = p.parse_args()
    print(json.dumps(PrecisionPDFGenerator(a.mapping_file).generate(a.excel_path)))
